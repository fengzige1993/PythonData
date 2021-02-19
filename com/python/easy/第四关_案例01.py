# 从openpyxl库导入load_workbook函数
from openpyxl import load_workbook

# 打开【10月员工绩效表】的工作簿，获取活动工作表
performance_wb  = load_workbook('./material/10月员工绩效表.xlsx')
performance_ws = performance_wb.active

# 打开【江宇工资信息表】的工作簿，获取活动工作表
info_wb = load_workbook('./material/江宇工资信息表.xlsx')
info_ws = info_wb.active

# 获取【绩效】值
performance = performance_ws['D14'].value
# 获取【奖金】值
bonus = performance_ws['E14'].value
# 获取【基本工资】值
base = performance_ws['F14'].value

# 写入【绩效】值
info_ws['E11'].value = performance
# 写入【奖金】值
info_ws['F11'].value = bonus
# 写入【基本工资】值
info_ws['G11'].value = base

# 保存对【江宇工资信息表】工作簿的写入
info_wb.save('./material/江宇工资信息表.xlsx')