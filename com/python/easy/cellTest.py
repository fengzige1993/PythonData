# 从openpyxl库导入load_workbook函数
from openpyxl  import load_workbook

# 打开【10月员工绩效表】的工作簿，获取活动工作表
performance_wb  = load_workbook('./10月员工绩效表.xlsx')
performance_ws = performance_wb.active

# 取出['A1']单元格对象
cell = performance_ws['A1']
# 使用单元格对象的value属性取出单元格的值，并打印
print(cell.value)